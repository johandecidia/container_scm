"""Tests for the Business Central PO XLSX layout parser."""

import io
from decimal import Decimal

import openpyxl
from django.test import SimpleTestCase

from apps.scm.imports.plugins.bc_po_xlsx.parser import (
    parse_bc_po_xlsx,
    to_flat_rows,
)
from apps.scm.imports.plugins.bc_po_xlsx.types import ParsedBCPO

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_bc_po_xlsx(
    po_number: str = "PO-2026-001",
    vendor_no: str = "SUPP-0001",
    vendor_name: str = "Acme Supplier Ltd",
    order_date: str = "2026-03-15",
    payment_terms: str = "Net 30",
    purchaser: str = "John Smith",
    currency: str = "USD",
    lines: list[dict] | None = None,
    layout: str = "en",  # "en" or "sv"
) -> io.BytesIO:
    """Build an in-memory XLSX file that mimics a BC PO document layout."""
    if lines is None:
        lines = [
            {"item_no": "ART-001", "description": "Widget A", "qty": 10, "uom": "PCS", "price": 25.0, "amount": 250.0},
            {"item_no": "ART-002", "description": "Widget B", "qty": 5, "uom": "PCS", "price": 100.0, "amount": 500.0},
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    if layout == "sv":
        # Swedish layout
        ws["A1"] = "Inköpsordernr."
        ws["B1"] = po_number
        ws["D1"] = "Orderdatum"
        ws["E1"] = order_date
        ws["A2"] = "Leverantörsnr."
        ws["B2"] = vendor_no
        ws["D2"] = "Betalningsvillkor"
        ws["E2"] = payment_terms
        ws["A3"] = "Leverantör"
        ws["B3"] = vendor_name
        ws["D3"] = "Inköpare"
        ws["E3"] = purchaser
        ws["A4"] = "Valutakod"
        ws["B4"] = currency
        ws["A5"] = ""
        ws["A6"] = "Nr."
        ws["B6"] = "Beskrivning"
        ws["C6"] = "Antal"
        ws["D6"] = "Enhet"
        ws["E6"] = "À-pris"
        ws["F6"] = "Belopp"
        for i, line in enumerate(lines, start=7):
            ws.cell(row=i, column=1, value=line["item_no"])
            ws.cell(row=i, column=2, value=line["description"])
            ws.cell(row=i, column=3, value=str(line["qty"]).replace(".", ","))
            ws.cell(row=i, column=4, value=line.get("uom", ""))
            ws.cell(row=i, column=5, value=str(line["price"]).replace(".", ","))
            ws.cell(row=i, column=6, value=str(line["amount"]).replace(".", ","))
        total_row = 7 + len(lines)
        ws.cell(row=total_row, column=1, value="Summa")
    else:
        # English layout
        ws["A1"] = "Vendor No."
        ws["B1"] = vendor_no
        ws["D1"] = "Order Date"
        ws["E1"] = order_date
        ws["A2"] = "Buy-from Vendor Name"
        ws["B2"] = vendor_name
        ws["D2"] = "Payment Terms"
        ws["E2"] = payment_terms
        ws["A3"] = "No."
        ws["B3"] = po_number
        ws["D3"] = "Purchaser"
        ws["E3"] = purchaser
        ws["A4"] = "Currency Code"
        ws["B4"] = currency
        ws["A5"] = ""
        ws["A6"] = "No."
        ws["B6"] = "Description"
        ws["C6"] = "Quantity"
        ws["D6"] = "Unit of Measure"
        ws["E6"] = "Direct Unit Cost"
        ws["F6"] = "Amount"
        for i, line in enumerate(lines, start=7):
            ws.cell(row=i, column=1, value=line["item_no"])
            ws.cell(row=i, column=2, value=line["description"])
            ws.cell(row=i, column=3, value=line["qty"])
            ws.cell(row=i, column=4, value=line.get("uom", ""))
            ws.cell(row=i, column=5, value=line["price"])
            ws.cell(row=i, column=6, value=line["amount"])
        total_row = 7 + len(lines)
        ws.cell(row=total_row, column=1, value="Total")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Parser tests
# ---------------------------------------------------------------------------


class TestBCPOXLSXParserEnglish(SimpleTestCase):
    """Test the parser against English-layout BC PO XLSX files."""

    def _parse(self, **kwargs) -> ParsedBCPO:
        xlsx = _make_bc_po_xlsx(**kwargs)
        result = parse_bc_po_xlsx(xlsx)
        self.assertIsNotNone(result)
        return result  # type: ignore[return-value]

    def test_extracts_po_number(self):
        result = self._parse(po_number="PO-2026-001")
        self.assertEqual(result.header.po_number, "PO-2026-001")

    def test_extracts_vendor_no(self):
        result = self._parse(vendor_no="SUPP-9999")
        self.assertEqual(result.header.vendor_no, "SUPP-9999")

    def test_extracts_vendor_name(self):
        result = self._parse(vendor_name="Global Parts AG")
        self.assertEqual(result.header.vendor_name, "Global Parts AG")

    def test_extracts_order_date(self):
        result = self._parse(order_date="2026-05-10")
        self.assertEqual(result.header.order_date, "2026-05-10")

    def test_extracts_payment_terms(self):
        result = self._parse(payment_terms="Net 60")
        self.assertEqual(result.header.payment_terms, "Net 60")

    def test_extracts_purchaser(self):
        result = self._parse(purchaser="Jane Doe")
        self.assertEqual(result.header.purchaser, "Jane Doe")

    def test_extracts_currency(self):
        result = self._parse(currency="SEK")
        self.assertEqual(result.header.currency, "SEK")

    def test_extracts_all_po_lines(self):
        result = self._parse(
            lines=[
                {"item_no": "A-001", "description": "Item 1", "qty": 10, "uom": "PCS", "price": 5.0, "amount": 50.0},
                {"item_no": "A-002", "description": "Item 2", "qty": 20, "uom": "PCS", "price": 3.0, "amount": 60.0},
                {"item_no": "A-003", "description": "Item 3", "qty": 1, "uom": "EA", "price": 200.0, "amount": 200.0},
            ]
        )
        self.assertEqual(len(result.lines), 3)
        self.assertEqual(result.lines[0].item_no, "A-001")
        self.assertEqual(result.lines[1].item_no, "A-002")
        self.assertEqual(result.lines[2].item_no, "A-003")

    def test_extracts_line_quantities(self):
        result = self._parse(
            lines=[
                {"item_no": "X-001", "description": "Part", "qty": 42, "uom": "PCS", "price": 1.5, "amount": 63.0},
            ]
        )
        self.assertEqual(result.lines[0].quantity, Decimal("42"))

    def test_extracts_line_unit_price(self):
        result = self._parse(
            lines=[
                {"item_no": "X-001", "description": "Part", "qty": 10, "uom": "PCS", "price": 99.99, "amount": 999.9},
            ]
        )
        self.assertEqual(result.lines[0].unit_price, Decimal("99.99"))

    def test_extracts_line_amount(self):
        result = self._parse(
            lines=[
                {"item_no": "X-001", "description": "Part", "qty": 10, "uom": "PCS", "price": 5.0, "amount": 50.0},
            ]
        )
        self.assertEqual(result.lines[0].amount, Decimal("50"))

    def test_line_source_row_number(self):
        """Lines track the source spreadsheet row for matching."""
        result = self._parse(
            lines=[
                {"item_no": "X-001", "description": "First", "qty": 1, "uom": "PCS", "price": 1.0, "amount": 1.0},
                {"item_no": "X-002", "description": "Second", "qty": 1, "uom": "PCS", "price": 1.0, "amount": 1.0},
            ]
        )
        self.assertEqual(len(result.lines), 2)
        # source_row is 1-based spreadsheet row number — second line must be after the first
        self.assertGreater(result.lines[1].source_row, result.lines[0].source_row)

    def test_handles_duplicate_item_numbers(self):
        """Same item_no on multiple lines must all be extracted (not collapsed)."""
        result = self._parse(
            lines=[
                {
                    "item_no": "SAME-ITEM",
                    "description": "First batch",
                    "qty": 5,
                    "uom": "PCS",
                    "price": 10.0,
                    "amount": 50.0,
                },
                {
                    "item_no": "SAME-ITEM",
                    "description": "Second batch",
                    "qty": 3,
                    "uom": "PCS",
                    "price": 10.0,
                    "amount": 30.0,
                },
                {
                    "item_no": "SAME-ITEM",
                    "description": "Third batch",
                    "qty": 2,
                    "uom": "PCS",
                    "price": 10.0,
                    "amount": 20.0,
                },
            ]
        )
        self.assertEqual(len(result.lines), 3)
        for line in result.lines:
            self.assertEqual(line.item_no, "SAME-ITEM")

    def test_amount_mismatch_produces_warning(self):
        """Lines where amount ≠ qty × unit_price generate a warning."""
        result = self._parse(
            lines=[
                {"item_no": "X-001", "description": "Part", "qty": 10, "uom": "PCS", "price": 5.0, "amount": 999.0},
            ]
        )
        self.assertTrue(
            any("amount" in w.lower() or "≠" in w for w in result.parse_warnings),
            f"Expected amount-mismatch warning but got: {result.parse_warnings}",
        )

    def test_stops_at_total_marker(self):
        """Parser stops extracting lines when it hits a 'Total' row."""
        result = self._parse(
            lines=[
                {"item_no": "A-001", "description": "Item 1", "qty": 5, "uom": "PCS", "price": 10.0, "amount": 50.0},
                {"item_no": "A-002", "description": "Item 2", "qty": 2, "uom": "PCS", "price": 25.0, "amount": 50.0},
            ]
        )
        # Must not include a "Total" row as a line item
        item_nos = [ln.item_no for ln in result.lines]
        self.assertNotIn("total", [n.lower() for n in item_nos])
        self.assertNotIn("Total", item_nos)

    def test_returns_none_for_missing_po_number(self):
        """A file without an identifiable PO number returns None."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Some random content"
        ws["A2"] = "No PO number label here"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = parse_bc_po_xlsx(buf)
        self.assertIsNone(result)

    def test_returns_none_for_empty_workbook(self):
        """An empty XLSX file returns None."""
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        result = parse_bc_po_xlsx(buf)
        self.assertIsNone(result)


class TestBCPOXLSXParserSwedish(SimpleTestCase):
    """Test the parser against Swedish-layout BC PO XLSX files."""

    def _parse(self, **kwargs) -> ParsedBCPO:
        kwargs.setdefault("layout", "sv")
        xlsx = _make_bc_po_xlsx(**kwargs)
        result = parse_bc_po_xlsx(xlsx)
        self.assertIsNotNone(result)
        return result  # type: ignore[return-value]

    def test_extracts_po_number_swedish(self):
        result = self._parse(po_number="IO-2026-042")
        self.assertEqual(result.header.po_number, "IO-2026-042")

    def test_extracts_vendor_no_swedish(self):
        result = self._parse(vendor_no="LEV-0099")
        self.assertEqual(result.header.vendor_no, "LEV-0099")

    def test_extracts_vendor_name_swedish(self):
        result = self._parse(vendor_name="Leverantören AB")
        self.assertEqual(result.header.vendor_name, "Leverantören AB")

    def test_extracts_order_date_swedish(self):
        result = self._parse(order_date="2026-01-20")
        self.assertEqual(result.header.order_date, "2026-01-20")

    def test_extracts_all_lines_swedish(self):
        result = self._parse(
            lines=[
                {"item_no": "P-100", "description": "Del A", "qty": 15, "uom": "ST", "price": 10, "amount": 150},
                {"item_no": "P-200", "description": "Del B", "qty": 8, "uom": "ST", "price": 25, "amount": 200},
            ]
        )
        self.assertEqual(len(result.lines), 2)

    def test_swedish_comma_decimal_quantity(self):
        """Swedish locale uses comma as decimal separator."""
        xlsx = _make_bc_po_xlsx(
            layout="sv",
            lines=[
                {"item_no": "X-100", "description": "Part", "qty": 1.5, "uom": "M", "price": 10, "amount": 15},
            ],
        )
        result = parse_bc_po_xlsx(xlsx)
        self.assertIsNotNone(result)
        # "1,5" should be parsed as 1.5
        self.assertEqual(result.lines[0].quantity, Decimal("1.5"))  # type: ignore[union-attr]


class TestToFlatRows(SimpleTestCase):
    """Test the flat row serialiser used by the import pipeline."""

    def _get_rows(self, **kwargs) -> list[dict]:
        xlsx = _make_bc_po_xlsx(**kwargs)
        parsed = parse_bc_po_xlsx(xlsx)
        self.assertIsNotNone(parsed)
        return to_flat_rows(parsed)  # type: ignore[arg-type]

    def test_one_row_per_line(self):
        rows = self._get_rows(
            lines=[
                {"item_no": "A", "description": "A", "qty": 1, "uom": "PCS", "price": 1, "amount": 1},
                {"item_no": "B", "description": "B", "qty": 2, "uom": "PCS", "price": 2, "amount": 4},
                {"item_no": "C", "description": "C", "qty": 3, "uom": "PCS", "price": 3, "amount": 9},
            ]
        )
        self.assertEqual(len(rows), 3)

    def test_header_fields_repeated_on_every_row(self):
        rows = self._get_rows(
            po_number="PO-TEST-1",
            vendor_no="V-001",
            vendor_name="Test Supplier",
            lines=[
                {"item_no": "A", "description": "A", "qty": 1, "uom": "PCS", "price": 1, "amount": 1},
                {"item_no": "B", "description": "B", "qty": 1, "uom": "PCS", "price": 1, "amount": 1},
            ],
        )
        for row in rows:
            self.assertEqual(row["po_number"], "PO-TEST-1")
            self.assertEqual(row["supplier_no"], "V-001")
            self.assertEqual(row["supplier_name"], "Test Supplier")

    def test_canonical_field_names(self):
        """Output uses the canonical field names expected by the PO import schema."""
        rows = self._get_rows()
        required = {"po_number", "supplier_no", "supplier_name", "order_date", "line_no", "item_no", "ordered_qty"}
        for row in rows:
            for field in required:
                self.assertIn(field, row, f"Field {field!r} missing from row {row}")

    def test_line_numbers_are_unique(self):
        rows = self._get_rows(
            lines=[
                {"item_no": "X", "description": "X", "qty": 1, "uom": "PCS", "price": 1, "amount": 1},
                {"item_no": "X", "description": "X", "qty": 1, "uom": "PCS", "price": 1, "amount": 1},
                {"item_no": "X", "description": "X", "qty": 1, "uom": "PCS", "price": 1, "amount": 1},
            ]
        )
        line_nos = [row["line_no"] for row in rows]
        self.assertEqual(len(line_nos), len(set(line_nos)), "Duplicate line_no values found")

    def test_duplicate_item_numbers_preserved(self):
        """Same item_no on multiple rows must be preserved (not collapsed)."""
        rows = self._get_rows(
            lines=[
                {"item_no": "DUP", "description": "First", "qty": 5, "uom": "PCS", "price": 10, "amount": 50},
                {"item_no": "DUP", "description": "Second", "qty": 3, "uom": "PCS", "price": 10, "amount": 30},
            ]
        )
        self.assertEqual(len(rows), 2)
        self.assertTrue(all(r["item_no"] == "DUP" for r in rows))

    def test_currency_defaults_to_eur_when_missing(self):
        xlsx = _make_bc_po_xlsx(currency="")
        parsed = parse_bc_po_xlsx(xlsx)
        # If currency is blank/missing, to_flat_rows should default to EUR
        rows = to_flat_rows(parsed)  # type: ignore[arg-type]
        for row in rows:
            self.assertEqual(row["currency"], "EUR")
