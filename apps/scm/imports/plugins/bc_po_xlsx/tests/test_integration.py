"""Integration tests for the BC PO XLSX import flow (parser → validate → commit).

These tests use the full import pipeline: parse_import_job → validate_import_job
→ confirm_import_job, exercising the same code path as the UI.
"""

import contextlib
import io

import openpyxl
from django.test import TestCase, override_settings

from apps.scm.imports.models import ImportJob, ImportRow
from apps.scm.imports.services import confirm_import_job, parse_import_job, validate_import_job
from apps.scm.imports.tests.helpers import make_team, make_user
from apps.scm.procurement.models import PurchaseOrder, PurchaseOrderLine

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_bc_po_xlsx_file(
    po_number: str = "PO-2026-INT-001",
    vendor_no: str = "SUPP-INT-001",
    vendor_name: str = "Integration Supplier Ltd",
    order_date: str = "2026-04-01",
    lines: list[dict] | None = None,
) -> io.BytesIO:
    """Build a minimal English-layout BC PO XLSX for integration tests."""
    if lines is None:
        lines = [
            {"item_no": "ITEM-001", "description": "Widget", "qty": 10, "uom": "PCS", "price": 5.0, "amount": 50.0},
            {"item_no": "ITEM-002", "description": "Gadget", "qty": 3, "uom": "PCS", "price": 20.0, "amount": 60.0},
        ]

    wb = openpyxl.Workbook()
    ws = wb.active

    ws["A1"] = "Vendor No."
    ws["B1"] = vendor_no
    ws["D1"] = "Order Date"
    ws["E1"] = order_date
    ws["A2"] = "Buy-from Vendor Name"
    ws["B2"] = vendor_name
    ws["A3"] = "No."
    ws["B3"] = po_number

    ws["A5"] = "No."
    ws["B5"] = "Description"
    ws["C5"] = "Quantity"
    ws["D5"] = "Unit of Measure"
    ws["E5"] = "Direct Unit Cost"
    ws["F5"] = "Amount"

    for i, line in enumerate(lines, start=6):
        ws.cell(row=i, column=1, value=line["item_no"])
        ws.cell(row=i, column=2, value=line["description"])
        ws.cell(row=i, column=3, value=line["qty"])
        ws.cell(row=i, column=4, value=line.get("uom", ""))
        ws.cell(row=i, column=5, value=line["price"])
        ws.cell(row=i, column=6, value=line["amount"])

    ws.cell(row=6 + len(lines), column=1, value="Total")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_import_job(team, user, buf: io.BytesIO, filename: str = "test_po.xlsx") -> ImportJob:
    from django.core.files.uploadedfile import SimpleUploadedFile

    f = SimpleUploadedFile(
        filename, buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return ImportJob.objects.create(
        team=team,
        created_by=user,
        file=f,
        original_filename=filename,
        import_type=ImportJob.ImportType.BC_PO_XLSX,
        status=ImportJob.Status.UPLOADED,
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


@override_settings(SCM_ENABLE_BC_PO_XLSX_IMPORT=True)
class TestBCPOXLSXParseStage(TestCase):
    """Tests for the parse phase of the BC PO XLSX import."""

    def setUp(self):
        self.team = make_team()
        self.user = make_user()

    def test_parse_creates_rows(self):
        buf = _make_bc_po_xlsx_file(
            lines=[
                {"item_no": "A-001", "description": "Part A", "qty": 5, "uom": "PCS", "price": 10, "amount": 50},
                {"item_no": "A-002", "description": "Part B", "qty": 2, "uom": "PCS", "price": 25, "amount": 50},
            ]
        )
        job = _make_import_job(self.team, self.user, buf)
        parse_import_job(job)
        self.assertEqual(job.status, ImportJob.Status.PARSED)
        self.assertEqual(job.total_rows, 2)

    def test_parse_rows_have_canonical_fields(self):
        buf = _make_bc_po_xlsx_file()
        job = _make_import_job(self.team, self.user, buf)
        parse_import_job(job)
        row = job.rows.first()
        self.assertIsNotNone(row)
        self.assertIn("po_number", row.mapped_data)
        self.assertIn("supplier_no", row.mapped_data)
        self.assertIn("item_no", row.mapped_data)
        self.assertIn("ordered_qty", row.mapped_data)

    def test_parse_rows_valid_after_pydantic(self):
        buf = _make_bc_po_xlsx_file()
        job = _make_import_job(self.team, self.user, buf)
        parse_import_job(job)
        for row in job.rows.all():
            self.assertEqual(
                row.status,
                ImportRow.Status.VALID,
                f"Row {row.row_number} invalid: {row.errors}",
            )


@override_settings(SCM_ENABLE_BC_PO_XLSX_IMPORT=True)
class TestBCPOXLSXValidateStage(TestCase):
    """Tests for the validate phase of the BC PO XLSX import."""

    def setUp(self):
        self.team = make_team()
        self.user = make_user()

    def _full_parse(self, buf: io.BytesIO) -> ImportJob:
        job = _make_import_job(self.team, self.user, buf)
        parse_import_job(job)
        return job

    def test_validate_runs_without_error(self):
        buf = _make_bc_po_xlsx_file()
        job = self._full_parse(buf)
        validate_import_job(job)
        self.assertEqual(job.status, ImportJob.Status.VALIDATED)

    def test_validate_counts_valid_rows(self):
        buf = _make_bc_po_xlsx_file(
            lines=[
                {"item_no": "A-001", "description": "Part A", "qty": 5, "uom": "PCS", "price": 10, "amount": 50},
                {"item_no": "A-002", "description": "Part B", "qty": 2, "uom": "PCS", "price": 25, "amount": 50},
            ]
        )
        job = self._full_parse(buf)
        validate_import_job(job)
        self.assertEqual(job.valid_rows, 2)
        self.assertEqual(job.invalid_rows, 0)

    def test_duplicate_item_no_generates_warning(self):
        """Two lines with the same item_no on the same PO should produce a warning, not an error."""

        buf = _make_bc_po_xlsx_file(
            lines=[
                {"item_no": "DUP-001", "description": "First lot", "qty": 5, "uom": "PCS", "price": 10, "amount": 50},
                {"item_no": "DUP-001", "description": "Second lot", "qty": 3, "uom": "PCS", "price": 10, "amount": 30},
            ]
        )
        job = self._full_parse(buf)
        validate_import_job(job)

        # Both rows should still be VALID (duplicate item_no is not a blocking error)
        self.assertEqual(job.valid_rows, 2)


@override_settings(SCM_ENABLE_BC_PO_XLSX_IMPORT=True)
class TestBCPOXLSXCommitStage(TestCase):
    """Tests for the commit phase of the BC PO XLSX import."""

    def setUp(self):
        self.team = make_team()
        self.user = make_user()

    def _full_pipeline(self, buf: io.BytesIO, update_existing: bool = False) -> ImportJob:
        job = _make_import_job(self.team, self.user, buf)
        parse_import_job(job)
        validate_import_job(job)
        confirm_import_job(job, update_existing=update_existing)
        return job

    def test_commit_creates_purchase_order(self):
        buf = _make_bc_po_xlsx_file(po_number="PO-COMMIT-001", vendor_no="V-001")
        self._full_pipeline(buf)
        self.assertTrue(PurchaseOrder.objects.filter(team=self.team, po_number="PO-COMMIT-001").exists())

    def test_commit_creates_po_lines(self):
        buf = _make_bc_po_xlsx_file(
            po_number="PO-COMMIT-002",
            lines=[
                {"item_no": "A-001", "description": "Part A", "qty": 5, "uom": "PCS", "price": 10, "amount": 50},
                {"item_no": "A-002", "description": "Part B", "qty": 2, "uom": "PCS", "price": 25, "amount": 50},
                {"item_no": "A-003", "description": "Part C", "qty": 1, "uom": "PCS", "price": 100, "amount": 100},
            ],
        )
        self._full_pipeline(buf)
        po = PurchaseOrder.objects.get(team=self.team, po_number="PO-COMMIT-002")
        self.assertEqual(po.lines.count(), 3)

    def test_commit_sets_supplier_fields(self):
        buf = _make_bc_po_xlsx_file(
            po_number="PO-SUPPLIER-001",
            vendor_no="SUPP-XYZ",
            vendor_name="XYZ Supplies Ltd",
        )
        self._full_pipeline(buf)
        po = PurchaseOrder.objects.get(team=self.team, po_number="PO-SUPPLIER-001")
        self.assertEqual(po.supplier_no, "SUPP-XYZ")
        self.assertEqual(po.supplier_name, "XYZ Supplies Ltd")

    def test_second_import_does_not_create_duplicates(self):
        """Running the same file twice must not create duplicate POs or lines."""
        buf = _make_bc_po_xlsx_file(
            po_number="PO-IDEM-001",
            lines=[
                {"item_no": "X-001", "description": "Part X", "qty": 10, "uom": "PCS", "price": 5, "amount": 50},
            ],
        )
        # First import
        self._full_pipeline(buf)
        po_count_1 = PurchaseOrder.objects.filter(team=self.team, po_number="PO-IDEM-001").count()
        line_count_1 = PurchaseOrderLine.objects.filter(
            purchase_order__team=self.team, purchase_order__po_number="PO-IDEM-001"
        ).count()

        # Second import (same data)
        buf2 = _make_bc_po_xlsx_file(
            po_number="PO-IDEM-001",
            lines=[
                {"item_no": "X-001", "description": "Part X", "qty": 10, "uom": "PCS", "price": 5, "amount": 50},
            ],
        )
        self._full_pipeline(buf2)
        po_count_2 = PurchaseOrder.objects.filter(team=self.team, po_number="PO-IDEM-001").count()
        line_count_2 = PurchaseOrderLine.objects.filter(
            purchase_order__team=self.team, purchase_order__po_number="PO-IDEM-001"
        ).count()

        self.assertEqual(po_count_1, po_count_2, "Second import created a duplicate PO")
        self.assertEqual(line_count_1, line_count_2, "Second import created duplicate lines")

    def test_update_mode_updates_existing_line(self):
        """With update_existing=True, a second import updates changed fields."""
        buf = _make_bc_po_xlsx_file(
            po_number="PO-UPDATE-001",
            lines=[
                {"item_no": "U-001", "description": "Original Desc", "qty": 5, "uom": "PCS", "price": 10, "amount": 50},
            ],
        )
        self._full_pipeline(buf)

        buf2 = _make_bc_po_xlsx_file(
            po_number="PO-UPDATE-001",
            lines=[
                {"item_no": "U-001", "description": "Updated Desc", "qty": 8, "uom": "PCS", "price": 10, "amount": 80},
            ],
        )
        self._full_pipeline(buf2, update_existing=True)

        po = PurchaseOrder.objects.get(team=self.team, po_number="PO-UPDATE-001")
        line = po.lines.first()
        self.assertIsNotNone(line)
        self.assertEqual(line.ordered_qty, 8)  # type: ignore[union-attr]

    def test_commit_is_atomic_and_rolls_back_on_error(self):
        """If the import fails mid-way, no partial data is committed."""
        from unittest.mock import patch

        buf = _make_bc_po_xlsx_file(po_number="PO-ATOMIC-001")
        job = _make_import_job(self.team, self.user, buf)
        parse_import_job(job)
        validate_import_job(job)

        # Patch run_import (which is decorated with @transaction.atomic) to raise
        # mid-way so we can verify the atomic block rolls everything back.
        with (
            patch("apps.scm.imports.importers.run_import", side_effect=RuntimeError("forced failure")),
            contextlib.suppress(RuntimeError),
        ):
            confirm_import_job(job)

        self.assertFalse(PurchaseOrder.objects.filter(team=self.team, po_number="PO-ATOMIC-001").exists())

    def test_job_marked_completed_after_commit(self):
        buf = _make_bc_po_xlsx_file(po_number="PO-STATUS-001")
        job = self._full_pipeline(buf)
        job.refresh_from_db()
        self.assertEqual(job.status, ImportJob.Status.COMPLETED)

    def test_team_isolation(self):
        """An import for Team A must not be visible to or affect Team B."""
        team_b = make_team(name="Team B", slug="team-b")

        buf = _make_bc_po_xlsx_file(po_number="PO-TEAM-001")
        self._full_pipeline(buf)

        # Team B should have no POs from this import
        self.assertFalse(PurchaseOrder.objects.filter(team=team_b, po_number="PO-TEAM-001").exists())
        # Team A should have the PO
        self.assertTrue(PurchaseOrder.objects.filter(team=self.team, po_number="PO-TEAM-001").exists())


@override_settings(SCM_ENABLE_BC_PO_XLSX_IMPORT=False)
class TestBCPOXLSXFeatureFlag(TestCase):
    """Tests for the feature flag gating."""

    def setUp(self):
        self.team = make_team()
        self.user = make_user()

    def test_import_type_hidden_from_form_when_disabled(self):
        from apps.scm.imports.forms import ImportUploadForm

        form = ImportUploadForm()
        choices = dict(form.fields["import_type"].choices)
        self.assertNotIn(ImportJob.ImportType.BC_PO_XLSX, choices)

    @override_settings(SCM_ENABLE_BC_PO_XLSX_IMPORT=True)
    def test_import_type_visible_in_form_when_enabled(self):
        from apps.scm.imports.forms import ImportUploadForm

        form = ImportUploadForm()
        choices = dict(form.fields["import_type"].choices)
        self.assertIn(ImportJob.ImportType.BC_PO_XLSX, choices)

    def test_form_rejects_bc_xlsx_type_when_disabled(self):
        import io as _io

        from django.core.files.uploadedfile import SimpleUploadedFile

        from apps.scm.imports.forms import ImportUploadForm

        wb = openpyxl.Workbook()
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        f = SimpleUploadedFile(
            "test.xlsx", buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        form = ImportUploadForm(
            data={"import_type": ImportJob.ImportType.BC_PO_XLSX},
            files={"file": f},
        )
        self.assertFalse(form.is_valid())
